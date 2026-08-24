"""Импорт данных из Excel (ТЗ-00 раздел 9, FR-SP-04, FR-TN-08).

Формат файла — один лист, первая строка заголовок, колонки по шаблону:
  A корпус · B код места · C тип места · D площадь · E ФИО · F ИНН ·
  G паспорт · H телефон · I сумма аренды · J дата начисления (пусто = общая) ·
  K срок оплаты (пусто = общий) · L начальный долг

Одна строка = одно место одного арендатора. Арендатор с несколькими местами
повторяется в нескольких строках (ИНН одинаковый, начальный долг указывается
один раз — в первой строке).

Пробный прогон (dry_run) выполняет все проверки и возвращает отчёт об ошибках
(строка, поле, описание), не записывая ничего в базу.
"""
import datetime
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from django.db import transaction

from apps.billing.models import Charge
from apps.billing.services import create_manual_charge
from apps.catalog.models import Building, Spot
from apps.core.money import q2
from apps.core.services import audit
from apps.tenants.models import Tenant, TenantSpot
from apps.tenants.services import assign_spot

HEADER_HINT = ['корпус', 'код места', 'тип', 'площадь', 'фио', 'инн',
               'паспорт', 'телефон', 'сумма', 'дата начисления', 'срок оплаты', 'долг']

SPOT_TYPE_MAP = {
    'контейнер': Spot.Type.CONTAINER,
    'павильон': Spot.Type.PAVILION,
    'прилавок': Spot.Type.COUNTER,
    'бутик': Spot.Type.BOUTIQUE,
}


@dataclass
class ImportReport:
    errors: list[dict] = field(default_factory=list)
    created_buildings: int = 0
    created_spots: int = 0
    created_tenants: int = 0
    updated_tenants: int = 0
    created_links: int = 0
    initial_debt_rows: int = 0
    total_initial_debt: Decimal = Decimal('0.00')

    @property
    def ok(self) -> bool:
        return not self.errors

    def error(self, row: int, column: str, message: str):
        self.errors.append({'row': row, 'field': column, 'message': message})


def _cell(row, index) -> str:
    value = row[index] if index < len(row) else None
    if value is None:
        return ''
    return str(value).strip()


def _parse_rows(file_obj) -> list[tuple[int, list]]:
    from openpyxl import load_workbook
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # заголовок
        if row is None or all(v in (None, '') for v in row):
            continue
        rows.append((i, list(row)))
    return rows


def run_import(file_obj, *, dry_run: bool, actor=None) -> ImportReport:
    """Импорт с проверкой данных. Выполняется дважды: пробный и окончательный (ТЗ-00 п. 9)."""
    report = ImportReport()
    try:
        rows = _parse_rows(file_obj)
    except Exception:
        report.error(0, 'файл', 'Не удалось прочитать файл. Ожидается XLSX по шаблону.')
        return report
    if not rows:
        report.error(0, 'файл', 'В файле нет строк с данными.')
        return report

    # --- Проверка (полная, независимо от режима) --------------------------------
    parsed = []
    seen_debt_by_inn: dict[str, Decimal] = {}
    for line_number, row in rows:
        building_name = _cell(row, 0)
        spot_code = _cell(row, 1)
        spot_type_raw = _cell(row, 2).lower()
        area_raw = _cell(row, 3).replace(',', '.')
        full_name = _cell(row, 4)
        inn = _cell(row, 5).split('.')[0]  # Excel любит превращать ИНН в число
        passport = _cell(row, 6)
        phone = _cell(row, 7)
        amount_raw = _cell(row, 8).replace(',', '.').replace(' ', '')
        billing_day_raw = _cell(row, 9).split('.')[0]
        payment_day_raw = _cell(row, 10).split('.')[0]
        debt_raw = _cell(row, 11).replace(',', '.').replace(' ', '')

        if not building_name:
            report.error(line_number, 'корпус', 'Не указан корпус')
        if not spot_code:
            report.error(line_number, 'код места', 'Не указан код места')
        if not full_name:
            report.error(line_number, 'ФИО', 'Не указано ФИО арендатора')
        if not inn:
            report.error(line_number, 'ИНН', 'Не указан ИНН')
        elif not inn.isdigit() or not (8 <= len(inn) <= 20):
            report.error(line_number, 'ИНН', f'ИНН «{inn}» должен состоять из 8–20 цифр')

        spot_type = SPOT_TYPE_MAP.get(spot_type_raw, Spot.Type.OTHER if spot_type_raw else Spot.Type.CONTAINER)

        area = None
        if area_raw:
            try:
                area = q2(Decimal(area_raw))
            except InvalidOperation:
                report.error(line_number, 'площадь', f'Неверное число «{area_raw}»')

        amount = None
        if not amount_raw:
            report.error(line_number, 'сумма', 'Не указана сумма аренды')
        else:
            try:
                amount = q2(Decimal(amount_raw))
                if amount < 0:
                    report.error(line_number, 'сумма', 'Сумма аренды не может быть отрицательной')
            except InvalidOperation:
                report.error(line_number, 'сумма', f'Неверное число «{amount_raw}»')

        billing_day = payment_day = None
        if billing_day_raw:
            if billing_day_raw.isdigit() and 1 <= int(billing_day_raw) <= 31:
                billing_day = int(billing_day_raw)
            else:
                report.error(line_number, 'дата начисления', f'Неверный день «{billing_day_raw}»')
        if payment_day_raw:
            if payment_day_raw.isdigit() and 1 <= int(payment_day_raw) <= 31:
                payment_day = int(payment_day_raw)
            else:
                report.error(line_number, 'срок оплаты', f'Неверный день «{payment_day_raw}»')
        if billing_day and payment_day and billing_day >= payment_day:
            report.error(line_number, 'дата начисления',
                         'Дата начисления должна быть раньше срока оплаты')

        debt = None
        if debt_raw:
            try:
                debt = q2(Decimal(debt_raw))
                if debt < 0:
                    report.error(line_number, 'долг', 'Начальный долг не может быть отрицательным')
            except InvalidOperation:
                report.error(line_number, 'долг', f'Неверное число «{debt_raw}»')
            if inn in seen_debt_by_inn and debt is not None and debt > 0:
                report.error(line_number, 'долг',
                             'Начальный долг для этого ИНН уже указан в другой строке')
            elif debt is not None:
                seen_debt_by_inn[inn] = debt

        # Дубли кода места внутри файла
        if spot_code and any(p['spot_code'] == spot_code for p in parsed):
            report.error(line_number, 'код места', f'Код «{spot_code}» повторяется в файле')

        # Место уже занято в системе другим арендатором
        existing = Spot.objects.filter(code=spot_code).first()
        if existing is not None:
            active_link = TenantSpot.objects.filter(
                spot=existing, is_active=True).select_related('tenant').first()
            if active_link and active_link.tenant.inn != inn:
                report.error(line_number, 'код места',
                             f'Место «{spot_code}» уже занято арендатором '
                             f'{active_link.tenant.full_name}')

        parsed.append({
            'line': line_number, 'building_name': building_name, 'spot_code': spot_code,
            'spot_type': spot_type, 'area': area, 'full_name': full_name, 'inn': inn,
            'passport': passport, 'phone': phone, 'amount': amount,
            'billing_day': billing_day, 'payment_day': payment_day, 'debt': debt,
        })

    report.total_initial_debt = sum(seen_debt_by_inn.values(), Decimal('0.00'))
    report.initial_debt_rows = len([v for v in seen_debt_by_inn.values() if v > 0])

    if report.errors or dry_run:
        # Пробный прогон: посчитать, что будет создано
        building_names = {p['building_name'] for p in parsed if p['building_name']}
        report.created_buildings = len([
            n for n in building_names if not Building.objects.filter(name=n).exists()])
        report.created_spots = len([
            p for p in parsed
            if p['spot_code'] and not Spot.objects.filter(code=p['spot_code']).exists()])
        inns = {p['inn'] for p in parsed if p['inn']}
        report.created_tenants = len([
            i for i in inns if not Tenant.objects.filter(inn=i).exists()])
        report.updated_tenants = len(inns) - report.created_tenants
        return report

    # --- Запись (одна транзакция на весь файл) ----------------------------------
    with transaction.atomic():
        buildings: dict[str, Building] = {}
        for p in parsed:
            name = p['building_name']
            if name not in buildings:
                building = Building.objects.filter(name=name).first()
                if building is None:
                    code = name[:20]
                    if Building.objects.filter(code=code).exists():
                        code = f'{code[:15]}-{len(buildings)}'
                    building = Building.objects.create(name=name, code=code)
                    report.created_buildings += 1
                buildings[name] = building

        tenants: dict[str, Tenant] = {}
        for p in parsed:
            inn = p['inn']
            if inn in tenants:
                tenant = tenants[inn]
            else:
                tenant = Tenant.objects.filter(inn=inn).first()
                if tenant is None:
                    tenant = Tenant.objects.create(
                        full_name=p['full_name'], inn=inn,
                        passport_number=p['passport'], phone=p['phone'],
                        billing_day=p['billing_day'], payment_day=p['payment_day'])
                    report.created_tenants += 1
                else:
                    tenant.full_name = p['full_name'] or tenant.full_name
                    tenant.passport_number = p['passport'] or tenant.passport_number
                    tenant.phone = p['phone'] or tenant.phone
                    if p['billing_day']:
                        tenant.billing_day = p['billing_day']
                    if p['payment_day']:
                        tenant.payment_day = p['payment_day']
                    tenant.save()
                    report.updated_tenants += 1
                tenants[inn] = tenant

            spot = Spot.objects.filter(code=p['spot_code']).first()
            if spot is None:
                spot = Spot.objects.create(
                    building=buildings[p['building_name']], code=p['spot_code'],
                    spot_type=p['spot_type'], area_sqm=p['area'])
                report.created_spots += 1

            if not TenantSpot.objects.filter(tenant=tenant, spot=spot, is_active=True).exists():
                assign_spot(tenant=tenant, spot=spot, monthly_amount=p['amount'], actor=actor)
                report.created_links += 1

        # Начальный долг — отдельное начисление source=initial (FR-TN-08)
        for inn, debt in seen_debt_by_inn.items():
            if debt > 0:
                create_manual_charge(
                    tenant=tenants[inn], amount=debt,
                    comment='Долг на начало работы Системы', actor=actor,
                    source=Charge.Source.INITIAL)

        audit(action='import_excel', model_name='Tenant', object_id='import',
              actor=actor, new_value={
                  'tenants_created': report.created_tenants,
                  'spots_created': report.created_spots,
                  'links_created': report.created_links,
                  'initial_debt_total': str(report.total_initial_debt)})
    return report


def template_workbook_response():
    """Шаблон файла импорта, который Исполнитель предоставляет Заказчику (ТЗ-00 п. 9)."""
    from apps.reports.services import excel_response
    header = ['Корпус', 'Код места', 'Тип места', 'Площадь м2', 'ФИО арендатора',
              'ИНН', 'Паспорт', 'Телефон', 'Сумма аренды', 'Дата начисления',
              'Срок оплаты', 'Долг на начало']
    example = ['А', 'А-01', 'контейнер', '12.5', 'Иванов Иван Иванович',
               '12345678901234', 'AN1234567', '+996700123456', '12000', '', '', '5000']
    return excel_response('import_template.xlsx', header, [example])
